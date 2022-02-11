from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from anonymizer import config
from anonymizer.utils.json import write_to_json


def make_template_rules_file(
        files: list,
        filename: str = 'rules.json'
) -> str:
    """
    Создает файл конфигурации
    """
    rules = _make_template_rules(files)
    rules_filepath = write_to_json(
        rules,
        config.BASE_DIR / filename,
    )
    return rules_filepath


def _make_template_rules(files):
    """
    Формирует структуру конфигурационного файла
    на основе документов Excel
    """

    rules = {}

    for file in files:

        filename = str(file)
        base_dir = str(config.BASE_DIR)
        if base_dir in filename:
            filename = filename.replace(base_dir, '', 1)

        rules[filename] = _get_excel_structure(file)

    return rules


def _get_excel_structure(excel_file_path:str):
    """
    Создает конфиг по структуре файла Excel 
    """

    rules_file = {}

    wb = load_workbook(excel_file_path)

    for sheet_name in wb.sheetnames:

        # TODO: Спрашивать о наличии заголовка
        # TODO: Определять автоматически
        sheet_structure = {
            'has_title': True,
        }

        for column in range(1, wb[sheet_name].max_column+1):
            # TODO: Определять автоматически тип колонки
            sheet_structure[get_column_letter(column)] = {
                'algorithm': 'just_copy',
                'params': None,
            }

        rules_file[sheet_name] = sheet_structure

    return rules_file
